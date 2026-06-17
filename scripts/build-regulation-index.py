from __future__ import annotations

import json
import re
import zipfile
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from xml.etree import ElementTree as ET

try:
    import openpyxl
except ModuleNotFoundError:
    openpyxl = None


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "法律法规" / "法规索引数据.xlsx"
OUTPUT_JSON = ROOT / "database" / "regulation-index.json"
OUTPUT_JS = ROOT / "database" / "regulation-index.js"

TYPE_LABELS = {
    "law": "国家/地区法规",
    "treaty": "国际条约",
    "guide": "监管指南",
}

TOPIC_LABELS = {
    "privacy": "数据保护与隐私合规",
    "security": "网络安全与韧性",
    "flow": "数据流通与价值释放",
    "platform": "数字市场与平台治理",
    "ai": "人工智能与新技术治理",
    "market": "产品与市场准入",
    "mobility": "交通管理与行业数据监管",
    "crossborder": "数据跨境传输与监管遵从",
}

JURISDICTIONS = [
    ("eu", "欧盟", ["欧洲议会", "欧盟理事会", "european parliament", "council of the european union", "european commission", "edpb", "wp29", "article 29"]),
    ("germany", "德国", ["德国", "german", "deutsch", "bundestag", "bundesrat", "bdsg"]),
    ("sweden", "瑞典", ["瑞典", "swedish", "sverige", "sfs"]),
    ("portugal", "葡萄牙", ["葡萄牙", "portugal", "portuguese"]),
    ("spain", "西班牙", ["西班牙", "spain", "spanish", "aepd", "boe-a"]),
    ("greece", "希腊", ["希腊", "greece", "greek", "hellenic", "ελλάδα", "νoμο"]),
    ("ireland", "爱尔兰", ["爱尔兰", "ireland", "irish"]),
    ("denmark", "丹麦", ["丹麦", "denmark", "danish", "lov nr", "bek nr"]),
    ("france", "法国", ["法国", "france", "french", "cnil", "décret", "loi"]),
    ("italy", "意大利", ["意大利", "italy", "italian", "garante", "legge"]),
    ("netherlands", "荷兰", ["荷兰", "netherlands", "dutch", "nederland", "wet "]),
    ("belgium", "比利时", ["比利时", "belgium", "belgian"]),
    ("poland", "波兰", ["波兰", "poland", "polish"]),
    ("estonia", "爱沙尼亚", ["爱沙尼亚", "estonia", "estonian", "küberturvalisuse"]),
    ("czechia", "捷克", ["捷克", "czech"]),
    ("hungary", "匈牙利", ["匈牙利", "hungary", "hungarian", "országgyűlés"]),
    ("austria", "奥地利", ["奥地利", "austria", "austrian"]),
    ("finland", "芬兰", ["芬兰", "finland", "finnish"]),
    ("luxembourg", "卢森堡", ["卢森堡", "luxembourg"]),
    ("slovakia", "斯洛伐克", ["斯洛伐克", "slovakia", "slovak"]),
    ("slovenia", "斯洛文尼亚", ["斯洛文尼亚", "slovenia", "slovene"]),
    ("croatia", "克罗地亚", ["克罗地亚", "croatia", "croatian"]),
    ("romania", "罗马尼亚", ["罗马尼亚", "romania", "romanian"]),
    ("bulgaria", "保加利亚", ["保加利亚", "bulgaria", "bulgarian"]),
    ("lithuania", "立陶宛", ["立陶宛", "lithuania", "lithuanian"]),
    ("latvia", "拉脱维亚", ["拉脱维亚", "latvia", "latvian"]),
    ("malta", "马耳他", ["马耳他", "malta", "maltese"]),
    ("cyprus", "塞浦路斯", ["塞浦路斯", "cyprus", "cypriot"]),
]


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return re.sub(r"\s+", " ", str(value)).strip()


def text_blob(row: dict[str, str]) -> str:
    return " ".join(row.values()).lower()


def includes_any(text: str, keywords: list[str]) -> bool:
    return any(keyword.lower() in text for keyword in keywords)


def infer_type(text: str) -> str:
    if includes_any(text, ["guideline", "guidelines", "wp243", "wp248", "指南", "准则", "意见"]):
        return "guide"
    if includes_any(text, ["convention", "treaty", "公约", "条约"]):
        return "treaty"
    return "law"


def infer_topic(text: str) -> str:
    topic_rules = [
        ("ai", ["artificial intelligence", " ai ", "ai act", "人工智能", "算法"]),
        ("security", ["cyber", "nis2", "nis directive", "network security", "information security", "incident", "网络安全", "信息安全", "安全事件", "韧性"]),
        ("platform", ["digital services act", "digital markets act", " dsa", " dma", "gatekeeper", "platform", "平台", "守门人", "数字服务", "数字市场"]),
        ("flow", ["data act", "data governance act", "data governance", "data sharing", "data access", "reuse", "数据治理", "数据流通", "数据共享", "数据访问", "数据开放"]),
        ("crossborder", ["cross-border", "outside the european union", "international transfer", "transfer of personal data", "跨境", "国际传输", "境外", "向欧盟以外"]),
        ("mobility", ["vehicle", "automotive", "transport", "交通", "车辆", "汽车", "出行"]),
        ("market", ["product safety", "consumer", "market surveillance", "competition", "产品", "消费者", "市场监管", "竞争"]),
        ("privacy", ["gdpr", "data protection", "personal data", "privacy", "数据保护", "个人数据", "隐私", "数据主体"]),
    ]
    for topic, keywords in topic_rules:
        if includes_any(text, keywords):
            return topic
    return "privacy"


def infer_jurisdiction(text: str) -> tuple[str, str]:
    for key, label, keywords in JURISDICTIONS:
        if includes_any(text, keywords):
            return key, label
    return "europe", "欧洲"


def infer_status(effective_date: str) -> str:
    if not effective_date:
        return "待确认"
    try:
        effective = date.fromisoformat(effective_date)
    except ValueError:
        return "待确认"
    return "现行有效" if effective <= date.today() else "即将生效"


def make_id(index: int, title: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")
    return f"reg-{index:03d}-{slug[:54].strip('-') or 'item'}"


def ns_tag(namespace: str, tag: str) -> str:
    return f"{{{namespace}}}{tag}"


def column_index(cell_ref: str) -> int:
    letters = re.match(r"[A-Z]+", cell_ref.upper())
    if not letters:
        return 0
    value = 0
    for char in letters.group(0):
        value = value * 26 + ord(char) - ord("A") + 1
    return value - 1


def excel_date(serial: str) -> str:
    try:
        return (datetime(1899, 12, 30) + timedelta(days=float(serial))).date().isoformat()
    except ValueError:
        return clean(serial)


def load_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    strings = []
    for si in root.findall("x:si", ns):
        parts = [node.text or "" for node in si.findall(".//x:t", ns)]
        strings.append("".join(parts))
    return strings


def load_date_style_indexes(archive: zipfile.ZipFile) -> set[int]:
    date_num_fmt_ids = {14, 15, 16, 17, 22, 27, 30, 36, 50, 57}
    try:
        root = ET.fromstring(archive.read("xl/styles.xml"))
    except KeyError:
        return set()

    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    for fmt in root.findall("x:numFmts/x:numFmt", ns):
        fmt_id = int(fmt.attrib.get("numFmtId", "0"))
        fmt_code = fmt.attrib.get("formatCode", "").lower()
        if any(token in fmt_code for token in ("yy", "mm", "dd", "年", "月", "日")):
            date_num_fmt_ids.add(fmt_id)

    date_styles = set()
    for index, xf in enumerate(root.findall("x:cellXfs/x:xf", ns)):
        try:
            num_fmt_id = int(xf.attrib.get("numFmtId", "0"))
        except ValueError:
            continue
        if num_fmt_id in date_num_fmt_ids:
            date_styles.add(index)
    return date_styles


def first_sheet_path(archive: zipfile.ZipFile) -> tuple[str, str]:
    doc_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    pkg_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    sheet = workbook.find(f".//{ns_tag(doc_ns, 'sheet')}")
    if sheet is None:
        raise ValueError("Workbook has no sheets.")

    sheet_name = sheet.attrib.get("name", "Sheet1")
    relation_id = sheet.attrib.get(ns_tag(rel_ns, "id"))
    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    for rel in relationships.findall(ns_tag(pkg_rel_ns, "Relationship")):
        if rel.attrib.get("Id") == relation_id:
            target = rel.attrib["Target"].lstrip("/")
            if not target.startswith("xl/"):
                target = "xl/" + target
            return sheet_name, target
    raise ValueError(f"Could not resolve first sheet relationship {relation_id}.")


def cell_value(cell: ET.Element, shared_strings: list[str], date_styles: set[int]) -> str:
    value_node = cell.find(ns_tag("http://schemas.openxmlformats.org/spreadsheetml/2006/main", "v"))
    inline_node = cell.find(ns_tag("http://schemas.openxmlformats.org/spreadsheetml/2006/main", "is"))
    raw = value_node.text if value_node is not None else ""
    cell_type = cell.attrib.get("t", "")

    if cell_type == "s":
        try:
            return clean(shared_strings[int(raw)])
        except (ValueError, IndexError):
            return ""
    if cell_type == "inlineStr" and inline_node is not None:
        text_parts = [node.text or "" for node in inline_node.findall(".//" + ns_tag("http://schemas.openxmlformats.org/spreadsheetml/2006/main", "t"))]
        return clean("".join(text_parts))

    style_index = int(cell.attrib.get("s", "0") or 0)
    if raw and style_index in date_styles:
        return excel_date(raw)
    return clean(raw)


def read_rows_with_stdlib() -> list[dict[str, str]]:
    with zipfile.ZipFile(SOURCE) as archive:
        sheet_name, sheet_path = first_sheet_path(archive)
        globals()["SOURCE_SHEET_NAME"] = sheet_name
        shared_strings = load_shared_strings(archive)
        date_styles = load_date_style_indexes(archive)
        sheet_root = ET.fromstring(archive.read(sheet_path))

    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    parsed_rows = []
    for row in sheet_root.findall(".//x:sheetData/x:row", ns):
        cells = {}
        for cell in row.findall("x:c", ns):
            ref = cell.attrib.get("r", "")
            cells[column_index(ref)] = cell_value(cell, shared_strings, date_styles)
        if cells:
            width = max(cells) + 1
            parsed_rows.append([cells.get(index, "") for index in range(width)])

    headers = parsed_rows[0]
    records = []
    for values in parsed_rows[1:]:
        record = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        if any(record.values()):
            records.append(record)
    return records


def read_rows_with_openpyxl() -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(SOURCE, data_only=True)
    globals()["SOURCE_SHEET_NAME"] = workbook.active.title
    sheet = workbook.active
    headers = [clean(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    rows = []
    for row_index in range(2, sheet.max_row + 1):
        raw = {
            headers[col - 1]: clean(sheet.cell(row_index, col).value)
            for col in range(1, sheet.max_column + 1)
        }
        if any(raw.values()):
            rows.append(raw)
    return rows


def read_records() -> list[dict]:
    rows = read_rows_with_openpyxl() if openpyxl else read_rows_with_stdlib()
    records = []

    for raw in rows:
        blob = text_blob(raw)
        type_id = infer_type(blob)
        topic_id = infer_topic(blob)
        region_id, region_label = infer_jurisdiction(blob)
        publish_date = raw.get("发布日期", "")
        effective_date = raw.get("生效日期", "")
        title = raw.get("法规标题", "")
        local_title = raw.get("中文标题", "")

        records.append({
            "id": make_id(len(records) + 1, title or local_title),
            "title": title,
            "localTitle": local_title,
            "referenceNo": raw.get("发文字号", ""),
            "publishDate": publish_date,
            "effectiveDate": effective_date,
            "issuer": raw.get("发布机构", ""),
            "summary": raw.get("中文摘要", ""),
            "englishSummary": raw.get("英文摘要", ""),
            "originalUrl": raw.get("原文", ""),
            "region": region_id,
            "regionLabel": region_label,
            "type": type_id,
            "typeLabel": TYPE_LABELS[type_id],
            "topic": topic_id,
            "topicLabel": TOPIC_LABELS[topic_id],
            "status": infer_status(effective_date),
            "date": publish_date or effective_date,
        })

    return records


def build_payload(records: list[dict]) -> dict:
    region_counts = Counter(record["region"] for record in records)
    label_by_region = {record["region"]: record["regionLabel"] for record in records}
    regions = [
        {"id": region, "label": label_by_region[region], "count": count}
        for region, count in region_counts.most_common()
    ]
    regions.sort(key=lambda item: (0 if item["id"] == "eu" else 1, item["label"]))

    return {
        "meta": {
            "source": "data/法律法规/法规索引数据.xlsx",
            "sheet": globals().get("SOURCE_SHEET_NAME", "Sheet2"),
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "total": len(records),
            "fields": ["法规标题", "中文标题", "发文字号", "发布日期", "生效日期", "发布机构", "中文摘要", "英文摘要", "原文"],
        },
        "regions": regions,
        "types": [{"id": key, "label": value} for key, value in TYPE_LABELS.items()],
        "topics": [{"id": key, "label": value} for key, value in TOPIC_LABELS.items()],
        "results": records,
    }


def main() -> None:
    if not SOURCE.exists():
        raise SystemExit(f"Missing source workbook: {SOURCE}")

    payload = build_payload(read_records())
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    json_text = json.dumps(payload, ensure_ascii=False, indent=2)
    OUTPUT_JSON.write_text(json_text + "\n", encoding="utf-8")
    OUTPUT_JS.write_text(
        "// Generated by scripts/build-regulation-index.py. Do not edit by hand.\n"
        "(function () {\n"
        f"  window.RegulationIndexDatabase = {json_text};\n"
        "})();\n",
        encoding="utf-8",
    )
    print(f"Generated {OUTPUT_JSON.relative_to(ROOT)} and {OUTPUT_JS.relative_to(ROOT)} with {payload['meta']['total']} records.")


if __name__ == "__main__":
    main()
